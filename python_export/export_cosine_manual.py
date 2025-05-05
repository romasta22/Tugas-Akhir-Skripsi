import mysql.connector
import pandas as pd
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import os

# Koneksi Database
db = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="dbs_wisata"
)
cursor = db.cursor()

# Ambil Data
cursor.execute("SELECT id_wisata, nama_wisata, id_kategori, deskripsi_processed FROM tbl_wisata")
data = cursor.fetchall()

# Mapping Kategori
kategori_mapping = {1: "Wisata Alam", 2: "Wisata Budaya", 3: "Wisata Kuliner", 4: "Akomodasi"}

# Membuat DataFrame
documents = pd.DataFrame(data, columns=["ID Wisata", "Nama Wisata", "ID Kategori", "Deskripsi Processed"])
documents["Kategori"] = documents["ID Kategori"].map(kategori_mapping)

# Konversi ke Matriks TF-IDF
from sklearn.feature_extraction.text import TfidfVectorizer
vectorizer = TfidfVectorizer()
tfidf_matrix = vectorizer.fit_transform(documents["Deskripsi Processed"])

# Menghitung Cosine Similarity
cosine_sim = cosine_similarity(tfidf_matrix)

# Menyimpan Hasil ke DataFrame
cosine_rows = []
for i in tqdm(range(len(documents)), desc="Menghitung Cosine Similarity", unit="data"):
    for j in range(len(documents)):
        if i != j:
            dot_product = np.dot(tfidf_matrix[i].toarray(), tfidf_matrix[j].toarray().T)[0][0]
            norm_i = np.linalg.norm(tfidf_matrix[i].toarray())
            norm_j = np.linalg.norm(tfidf_matrix[j].toarray())
            cosine_score = cosine_sim[i][j]
            cosine_rows.append([
                documents["ID Wisata"][i],
                documents["Nama Wisata"][i],
                documents["Kategori"][i],
                documents["ID Wisata"][j],
                documents["Nama Wisata"][j],
                documents["Kategori"][j],
                dot_product,
                norm_i,
                norm_j,
                cosine_score
            ])

cosine_df = pd.DataFrame(cosine_rows, columns=["ID Wisata", "Nama Wisata", "Kategori Wisata", "ID Wisata Rekomendasi", "Nama Wisata Rekomendasi", "Kategori Rekomendasi", "Dot Product", "Norm 1", "Norm 2", "Nilai Cosine Similarity"])

# Nama File Baru
filename = "Coding_Sheet_Cosine_Manual.xlsx"

# Buat file Excel baru jika belum ada
if not os.path.exists(filename):
    wb = Workbook()
    wb.save(filename)
    print(f"File {filename} berhasil dibuat!")

wb = load_workbook(filename)
sheet = wb.create_sheet("Cosine Similarity Manual")

# Tambahkan Data
for r in dataframe_to_rows(cosine_df, index=False, header=True):
    sheet.append(r)

# Tambahkan Filter Otomatis
sheet.auto_filter.ref = sheet.dimensions

# Warna Otomatis untuk Nilai Cosine Similarity Terbesar
max_value = cosine_df["Nilai Cosine Similarity"].max()
fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=10, max_col=10):
    for cell in row:
        if cell.value == max_value:
            cell.fill = fill

# Simpan Kembali ke File Excel
wb.save(filename)

print(f"Export Cosine Similarity Berhasil ke {filename}!")
