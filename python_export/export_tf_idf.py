import mysql.connector
import pandas as pd
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
import string
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.styles import PatternFill
import math

# Koneksi Database
db = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="dbs_wisata"
)

cursor = db.cursor()

# Ambil data deskripsi dan kategori
cursor.execute("SELECT w.id_wisata, w.nama_wisata, w.deskripsi, w.id_kategori FROM tbl_wisata w")
data = cursor.fetchall()

# Mapping ID Kategori ke Nama Kategori
kategori_mapping = {
    1: "Wisata Alam",
    2: "Wisata Budaya",
    3: "Wisata Kuliner",
    4: "Akomodasi"
}

# Inisialisasi Stemming
factory = StemmerFactory()
stemmer = factory.create_stemmer()
stop_words = set(stopwords.words('indonesian'))

rows = []
terms = {}
all_tokens = []

print("Memproses data...")

# Proses Tokenizing & Simpan Semua Token
for row in tqdm(data, desc="Progress", unit="data"):
    id_wisata, nama, deskripsi, id_kategori = row
    kategori = kategori_mapping.get(id_kategori, "Tidak Diketahui")
    lower = deskripsi.lower()
    tokens = word_tokenize(lower)
    filtered = [word for word in tokens if word not in stop_words and word not in string.punctuation]
    stemming = [stemmer.stem(word) for word in filtered]
    all_tokens.append(stemming)
    rows.append([id_wisata, nama, kategori, stemming])

# Hitung DF (Document Frequency)
df = {}
for token_list in all_tokens:
    unique_tokens = set(token_list)
    for token in unique_tokens:
        if token in df:
            df[token] += 1
        else:
            df[token] = 1

# Hitung TF, IDF, dan TF-IDF
final_rows = []
total_documents = len(all_tokens)

for id_wisata, nama, kategori, tokens in tqdm(rows, desc="Menghitung TF-IDF", unit="data"):
    tf_dict = {}
    for token in tokens:
        tf_dict[token] = tf_dict.get(token, 0) + 1
    tf = {k: v / len(tokens) for k, v in tf_dict.items()}
    idf = {k: math.log(total_documents / (df[k] + 1)) for k in tf.keys()}  # Ditambahkan +1 untuk smoothing agar tidak error
    tf_idf = {k: tf[k] * idf[k] for k in tf.keys()}
    for token in tf.keys():
        final_rows.append([id_wisata, nama, kategori, token, tf[token], df[token], idf[token], tf[token] * idf[token], tf_idf[token]])

# Buat DataFrame Pandas
df = pd.DataFrame(final_rows, columns=["ID Wisata", "Nama Wisata", "Kategori", "Tokenizing", "TF", "DF", "IDF", "TF x IDF", "TF-IDF (Hasil Akhir)"])

# Ekspor ke Excel
filename = "Coding_Sheet_TF_IDF.xlsx"
df.to_excel(filename, index=False)

# Tambahkan Auto Filter + Warna Otomatis
wb = load_workbook(filename)
sheet = wb.active
sheet.auto_filter.ref = sheet.dimensions

max_value = df["TF-IDF (Hasil Akhir)"].max()
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=9, max_col=9):
    for cell in row:
        if cell.value == max_value:
            cell.fill = fill

wb.save(filename)

print("Export Berhasil dengan Auto Filter, Warna Otomatis, dan Perhitungan Manual!")
